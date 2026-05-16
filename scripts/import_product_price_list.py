"""Import product price list (.xlsx) into PostgreSQL `products`.

Price list (first worksheet, row 1 = headers):
  - Product name, salt/generic name, manufacturing company, expiry date,
    USD price per strip

Optional Schedule workbook (`--schedule-xlsx`): first sheet with columns for
Schedule X, Schedule H, and Schedule H1 (headers detected on row 1). Each
non-empty cell is a drug/salt term. If a term matches product name + salt name,
`schedule_category` is set to X, H1, or H (priority X > H1 > H when multiple match)
and `is_restricted` is True.

Upsert: rows with the same (product_name, salt_name, manufacturing_company, expiry_date)
after strip are updated on re-import. Multiple Excel lines for the same product with
different expiry dates are separate catalog rows (batch lines).

After import, orphan DB rows with null expiry_date but the same product triple as a
row that has an expiry are removed (leftovers from older imports).

Usage (from `wasa/`):
  alembic upgrade head
  python -m scripts.import_product_price_list path/to/prices.xlsx --schedule-xlsx path/to/schedule_hx.xlsx --dry-run
  python -m scripts.import_product_price_list path/to/prices.xlsx --schedule-xlsx path/to/schedule_hx.xlsx

See docs/PRODUCT_IMPORT.md.
"""

from __future__ import annotations

import argparse
import re
import sys
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path

from dotenv import load_dotenv
from openpyxl import load_workbook
from openpyxl.utils.datetime import from_excel

load_dotenv()

SCHEDULE_CATEGORIES: tuple[str, ...] = ("X", "H1", "H")
"""Match priority when a product hits terms in more than one schedule column."""

_SCHEDULE_HEADER_HINTS: tuple[tuple[str, str], ...] = (
    ("schedule h1", "H1"),
    ("schedule x", "X"),
    ("schedule h", "H"),
)

_HEADER_ALIASES: dict[str, tuple[str, ...]] = {
    "product_name": ("product name", "product", "item name", "trade name"),
    "salt_name": (
        "saltname",
        "salt name",
        "generic name",
        "salt",
        "generic",
        "salt/generic",
    ),
    "manufacturing_company": (
        "manufacturing company",
        "manufacturer",
        "mfg",
        "mfg company",
        "company",
    ),
    "expiry_date": ("expiry date", "expiry", "exp date", "best before"),
    "price_per_strip": (
        "usd price per strip",
        "price per strip",
        "usd price",
        "price",
        "unit price",
        "strip price",
    ),
}


def _norm(s: str) -> str:
    return re.sub(r"\s+", " ", (s or "").strip().lower())


def _resolve_headers(header_row: list[str | None]) -> dict[str, int]:
    normalized = [_norm(str(c) if c is not None else "") for c in header_row]
    resolved: dict[str, int] = {}
    for col_idx, cell in enumerate(normalized):
        if not cell:
            continue
        for field, aliases in _HEADER_ALIASES.items():
            if field in resolved:
                continue
            for alias in aliases:
                if cell == alias or alias in cell or cell in alias:
                    resolved[field] = col_idx
                    break
    return resolved


def _parse_decimal(raw: object) -> Decimal | None:
    if raw is None:
        return None
    if isinstance(raw, (int, float)):
        return Decimal(str(raw))
    s = str(raw).strip()
    if not s:
        return None
    s = re.sub(r"[^\d.,\-]", "", s).replace(",", "")
    if not s:
        return None
    try:
        return Decimal(s)
    except InvalidOperation:
        return None


_DATE_STRING_FORMATS = (
    "%Y-%m-%d",
    "%Y-%m-%d %H:%M:%S",
    "%Y-%m-%d %H:%M",
    "%Y-%m-%d %H:%M:%S.%f",
    "%d/%m/%Y",
    "%m/%d/%Y",
    "%d-%m-%Y",
    "%d-%m-%Y %H:%M:%S",
    "%d-%b-%Y",
    "%d %b %Y",
    "%d/%m/%y",
    "%d-%m-%y",
)


def _parse_date(raw: object) -> date | None:
    """Parse Excel expiry cells: datetime/date objects, serial numbers, or text."""
    if raw is None:
        return None
    if isinstance(raw, datetime):
        return raw.date()
    if isinstance(raw, date):
        return raw
    if isinstance(raw, (int, float)):
        try:
            converted = from_excel(raw)
            if isinstance(converted, datetime):
                return converted.date()
            if isinstance(converted, date):
                return converted
        except (ValueError, TypeError, OverflowError):
            return None
        return None

    s = str(raw).strip()
    if not s:
        return None
    for fmt in _DATE_STRING_FORMATS:
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    # e.g. "2028-01-01 00:00:00.000000" from str(datetime)
    head = s.split()[0] if " " in s else s
    for fmt in ("%Y-%m-%d", "%d-%m-%Y", "%d/%m/%Y", "%m/%d/%Y"):
        try:
            return datetime.strptime(head, fmt).date()
        except ValueError:
            continue
    return None


def _cell_raw(row: tuple[object, ...], colmap: dict[str, int], field: str) -> object | None:
    idx = colmap.get(field)
    if idx is None or idx >= len(row):
        return None
    return row[idx]


def _expiry_cell_has_value(raw: object | None) -> bool:
    if raw is None:
        return False
    if isinstance(raw, str):
        return bool(raw.strip())
    return True


def _log_expiry_parse_issues(issues: list[tuple[int, str, object, str]]) -> None:
    if not issues:
        return
    print(
        f"WARNING: {len(issues)} row(s) have an expiry value that could not be parsed:",
        file=sys.stderr,
    )
    for row_num, product_name, raw, raw_type in issues:
        print(
            f"  row {row_num}: {product_name!r}  raw={raw!r}  type={raw_type}",
            file=sys.stderr,
        )


def _category_from_schedule_header(cell: object) -> str | None:
    text = _norm(str(cell) if cell is not None else "").rstrip(":")
    if not text:
        return None
    for hint, category in _SCHEDULE_HEADER_HINTS:
        if hint in text:
            return category
    return None


def load_schedule_terms_by_category(path: Path) -> dict[str, set[str]]:
    """Load Schedule X / H / H1 drug terms from all columns on the first sheet."""
    if not path.is_file():
        raise FileNotFoundError(path)
    by_category: dict[str, set[str]] = {cat: set() for cat in SCHEDULE_CATEGORIES}
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return by_category

        col_to_category: dict[int, str] = {}
        for col_idx, cell in enumerate(rows[0]):
            cat = _category_from_schedule_header(cell)
            if cat:
                col_to_category[col_idx] = cat

        if not col_to_category and len(rows[0]) >= 3:
            col_to_category = {0: "X", 1: "H", 2: "H1"}

        for row in rows[1:]:
            if not row:
                continue
            for col_idx, category in col_to_category.items():
                if col_idx >= len(row):
                    continue
                v = row[col_idx]
                if v is None:
                    continue
                t = _norm(str(v))
                if len(t) >= 3:
                    by_category[category].add(t)
    finally:
        wb.close()
    return by_category


def _match_schedule_category(
    product_name: str,
    salt_name: str | None,
    schedule_by_category: dict[str, set[str]],
) -> str | None:
    if not any(schedule_by_category.values()):
        return None
    hay = _norm(product_name) + " " + _norm(salt_name or "")
    matched: set[str] = set()
    for category, terms in schedule_by_category.items():
        for term in sorted(terms, key=len, reverse=True):
            if len(term) < 3:
                continue
            if term in hay:
                matched.add(category)
                break
    for category in SCHEDULE_CATEGORIES:
        if category in matched:
            return category
    return None


def _row_to_payload(
    row: tuple[object, ...],
    colmap: dict[str, int],
    row_num: int,
    schedule_by_category: dict[str, set[str]],
    *,
    expiry_parse_issues: list[tuple[int, str, object, str]] | None = None,
) -> dict[str, object] | None:
    def cell(field: str) -> str:
        idx = colmap.get(field)
        if idx is None or idx >= len(row):
            return ""
        v = row[idx]
        if v is None:
            return ""
        return str(v).strip()

    pname = cell("product_name")
    if not pname:
        return None

    salt = cell("salt_name") or None
    mfg = cell("manufacturing_company") or None
    price = _parse_decimal(cell("price_per_strip"))
    if price is None or price <= 0:
        print(f"  SKIP row {row_num}: missing or invalid price for {pname!r}", file=sys.stderr)
        return None

    expiry_raw: object | None = None
    exp: date | None = None
    if "expiry_date" in colmap:
        expiry_raw = _cell_raw(row, colmap, "expiry_date")
        exp = _parse_date(expiry_raw)
        if exp is None and _expiry_cell_has_value(expiry_raw) and expiry_parse_issues is not None:
            expiry_parse_issues.append(
                (row_num, pname, expiry_raw, type(expiry_raw).__name__),
            )

    schedule_category = _match_schedule_category(pname, salt, schedule_by_category)

    return {
        "product_name": pname,
        "salt_name": salt,
        "manufacturing_company": mfg,
        "expiry_date": exp,
        "price_per_strip": price,
        "is_restricted": schedule_category is not None,
        "schedule_category": schedule_category,
    }


def _same_optional_str(a: str | None, b: str | None) -> bool:
    return (a or None) == (b or None)


def _same_expiry(a: date | None, b: date | None) -> bool:
    return (a or None) == (b or None)


def _find_existing(db, payload: dict[str, object]):
    """Match catalog row including expiry (same product can have multiple batch lines)."""
    from app.db.models import Product

    sn = payload.get("salt_name")
    mg = payload.get("manufacturing_company")
    exp = payload.get("expiry_date")
    q = db.query(Product).filter(Product.product_name == payload["product_name"])
    for row in q.all():
        if (
            _same_optional_str(row.salt_name, sn)  # type: ignore[arg-type]
            and _same_optional_str(row.manufacturing_company, mg)  # type: ignore[arg-type]
            and _same_expiry(row.expiry_date, exp)  # type: ignore[arg-type]
        ):
            return row
    return None


def _prune_orphan_null_expiry_duplicates(db) -> int:
    """Remove stale rows: null expiry but same product triple exists with a real expiry."""
    from app.db.models import Product

    pruned = 0
    null_rows = db.query(Product).filter(Product.expiry_date.is_(None)).all()
    for row in null_rows:
        sibling = (
            db.query(Product)
            .filter(
                Product.id != row.id,
                Product.product_name == row.product_name,
                Product.expiry_date.isnot(None),
            )
            .all()
        )
        for other in sibling:
            if _same_optional_str(row.salt_name, other.salt_name) and _same_optional_str(
                row.manufacturing_company, other.manufacturing_company
            ):
                db.delete(row)
                pruned += 1
                break
    return pruned


def import_workbook(
    path: Path,
    *,
    dry_run: bool,
    schedule_path: Path | None,
) -> tuple[int, int, int, dict[str, int], int]:
    if not path.is_file():
        raise FileNotFoundError(path)

    schedule_by_category: dict[str, set[str]] = {cat: set() for cat in SCHEDULE_CATEGORIES}
    if schedule_path:
        schedule_by_category = load_schedule_terms_by_category(schedule_path)
        print(
            f"Loaded schedule terms from {schedule_path}: "
            f"X={len(schedule_by_category['X'])} "
            f"H={len(schedule_by_category['H'])} "
            f"H1={len(schedule_by_category['H1'])}"
        )

    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb.active
        rows = ws.iter_rows(values_only=True)
        header = next(rows, None)
        if not header:
            raise ValueError("Empty workbook")

        colmap = _resolve_headers(list(header))
        required = ("product_name", "price_per_strip")
        missing = [f for f in required if f not in colmap]
        if missing:
            raise ValueError(
                f"Could not resolve columns {missing}. Found headers: {list(header)!r}. "
                f"See _HEADER_ALIASES in {__file__}"
            )

        inserted = updated = skipped = 0
        pending: list[dict[str, object]] = []
        expiry_parse_issues: list[tuple[int, str, object, str]] = []
        expiry_parsed = expiry_missing = 0

        for row_num, row in enumerate(rows, start=2):
            if not row or all(v is None or str(v).strip() == "" for v in row):
                skipped += 1
                continue
            payload = _row_to_payload(
                tuple(row),
                colmap,
                row_num,
                schedule_by_category,
                expiry_parse_issues=expiry_parse_issues,
            )
            if payload is None:
                skipped += 1
                continue
            pending.append(payload)
            if payload.get("expiry_date") is not None:
                expiry_parsed += 1
            elif "expiry_date" in colmap:
                expiry_missing += 1

        _log_expiry_parse_issues(expiry_parse_issues)
        expiry_stats = {
            "parsed": expiry_parsed,
            "missing": expiry_missing,
            "unparsed": len(expiry_parse_issues),
        }

        if dry_run:
            print(f"--dry-run: parsed {len(pending)} product rows (skipped {skipped} empty/invalid)")
            print(
                f"  expiry: parsed={expiry_parsed} empty_in_sheet={expiry_missing} "
                f"unparsed={len(expiry_parse_issues)}"
            )
            for p in pending[:5]:
                r = p.get("schedule_category") or "—"
                exp = p.get("expiry_date")
                print(
                    f"  {p['product_name']!r}  price={p['price_per_strip']}  "
                    f"expiry={exp}  schedule={r}"
                )
            if len(pending) > 5:
                print(f"  ... and {len(pending) - 5} more")
            return inserted, updated, skipped, expiry_stats, 0

        from sqlalchemy.orm import Session

        from app.db.database import SessionLocal
        from app.db.models import Product

        db: Session = SessionLocal()
        pruned = 0
        try:
            for payload in pending:
                existing = _find_existing(db, payload)
                if existing is None:
                    db.add(Product(**payload))  # type: ignore[arg-type]
                    inserted += 1
                else:
                    for k, v in payload.items():
                        setattr(existing, k, v)
                    updated += 1
            pruned = _prune_orphan_null_expiry_duplicates(db)
            db.commit()
        except Exception:
            db.rollback()
            raise
        finally:
            db.close()

        return inserted, updated, skipped, expiry_stats, pruned
    finally:
        wb.close()


def main() -> None:
    parser = argparse.ArgumentParser(description="Import product price list into products table.")
    parser.add_argument("xlsx", type=Path, help="Path to catalog .xlsx")
    parser.add_argument(
        "--schedule-xlsx",
        type=Path,
        default=None,
        help="Optional workbook: Schedule X / H / H1 columns on first sheet",
    )
    parser.add_argument("--dry-run", action="store_true", help="Parse only; no DB writes")
    args = parser.parse_args()

    try:
        ins, upd, sk, expiry_stats, pruned = import_workbook(
            args.xlsx,
            dry_run=args.dry_run,
            schedule_path=args.schedule_xlsx,
        )
    except Exception as exc:
        print(f"ERROR: {exc}", file=sys.stderr)
        sys.exit(1)

    if not args.dry_run:
        print(f"Done. inserted={ins} updated={upd} skipped_empty_or_bad={sk} pruned_orphans={pruned}")
        print(
            f"Expiry: parsed={expiry_stats['parsed']} empty_in_sheet={expiry_stats['missing']} "
            f"unparsed={expiry_stats['unparsed']}"
        )


if __name__ == "__main__":
    main()
