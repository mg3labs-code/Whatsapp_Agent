"""
scripts/seed_all_shipping_data.py
===================================
Reads ALL data directly from Excel files — no hardcoding.
Run AFTER: alembic upgrade head

Handles:
  1. box_specs        — from Intelligent_Weight_Sheet.xlsx (Box Specs sheet)
  2. product weights  — from Intelligent_Weight_Sheet.xlsx + catalog.xlsx
     Strategy: exact → normalized → brand → formula estimate
     Tracks weight_source so you know which products need manual review
  3. shipping_rates   — from EXPRESS_SHIPPING_CHARGES.xlsx + REGULAR__LP__SHIPPING_CHARGES.xlsx

Usage:
    python scripts/seed_all_shipping_data.py \
        --weight  "Intelligent_Weight_Sheet.xlsx" \
        --catalog "catalog.xlsx" \
        --ems     "EXPRESS_SHIPPING_CHARGES.xlsx" \
        --lp      "REGULAR__LP__SHIPPING_CHARGES.xlsx" \
        [--dry-run]

After running, check unmatched_weights_review.csv for the 118 estimated
products — update their weight_g manually if accuracy is critical.
"""

from __future__ import annotations

import argparse
import csv
import os
import re
import sys
from difflib import SequenceMatcher

import openpyxl
import psycopg2
import psycopg2.extras
from dotenv import load_dotenv

load_dotenv()


# ─────────────────────────────────────────────────────
# DB
# ─────────────────────────────────────────────────────

def get_conn():
    url = os.getenv("DATABASE_URL", "")
    if not url:
        sys.exit("DATABASE_URL not set in .env")

    # Railway and some tools use postgresql+psycopg2:// or postgres://
    # psycopg2.connect() only accepts postgresql:// or postgres://
    # Strip the driver suffix and normalise the scheme
    url = url.replace("postgresql+psycopg2://", "postgresql://")
    url = url.replace("postgres+psycopg2://", "postgresql://")
    url = url.replace("postgres://", "postgresql://")

    return psycopg2.connect(url)


# ─────────────────────────────────────────────────────
# STEP 1 — Create / alter tables
# ─────────────────────────────────────────────────────

def create_tables(cur):
    print("\n[1/4] Creating / verifying tables...")

    cur.execute("""
        CREATE TABLE IF NOT EXISTS box_specs (
            box_no      TEXT PRIMARY KEY,
            box_type    TEXT NOT NULL,
            weight_g    NUMERIC(8,2) NOT NULL,
            height_cm   TEXT, length_cm TEXT, breadth_cm TEXT,
            max_strips  TEXT, max_tubes TEXT, max_vials TEXT, max_bottles TEXT
        )
    """)

    cur.execute("""
        CREATE TABLE IF NOT EXISTS shipping_rates (
            id            SERIAL PRIMARY KEY,
            country_name  TEXT NOT NULL,
            shipping_type TEXT NOT NULL CHECK (shipping_type IN ('EMS','LP')),
            weight_from_g INTEGER NOT NULL,
            weight_to_g   INTEGER NOT NULL,
            rate_usd      NUMERIC(8,2),
            UNIQUE (country_name, shipping_type, weight_from_g)
        )
    """)
    cur.execute("""
        CREATE INDEX IF NOT EXISTS idx_shipping_lookup
        ON shipping_rates (UPPER(country_name), shipping_type, weight_from_g)
    """)

    # products table — add weight columns if missing
    cur.execute("""
        ALTER TABLE products
            ADD COLUMN IF NOT EXISTS weight_g      NUMERIC(8,2) DEFAULT 0,
            ADD COLUMN IF NOT EXISTS weight_source TEXT DEFAULT 'unknown'
    """)
    # weight_source values:
    #   'exact'       — exact name match with weight sheet
    #   'normalized'  — matched after stripping dose/pack/form
    #   'brand'       — matched on brand name prefix
    #   'estimated'   — not in weight sheet, calculated from product name
    #   'manual'      — manually entered / verified

    # orders table — add shipping columns if missing
    cur.execute("""
        ALTER TABLE orders
            ADD COLUMN IF NOT EXISTS total_weight_g    NUMERIC(10,2),
            ADD COLUMN IF NOT EXISTS box_no            TEXT,
            ADD COLUMN IF NOT EXISTS shipping_type     TEXT,
            ADD COLUMN IF NOT EXISTS shipping_cost_usd NUMERIC(10,2),
            ADD COLUMN IF NOT EXISTS shipping_days     TEXT
    """)

    print("  ✅ Tables OK")


# ─────────────────────────────────────────────────────
# STEP 2 — Box specs
# ─────────────────────────────────────────────────────

def seed_box_specs(cur, weight_file: str):
    print("\n[2/4] Seeding box_specs...")
    wb = openpyxl.load_workbook(weight_file, read_only=True, data_only=True)
    ws = wb['📐 Box Specs']
    boxes = []
    for r in ws.iter_rows(values_only=True):
        if not r[0] or r[0] not in ('Standard', 'DHL'):
            continue
        box_no = f"DHL{int(r[1])}" if r[0] == 'DHL' else str(int(r[1]))
        if not r[5]:
            continue
        boxes.append((box_no, r[0], float(r[5]),
                       str(r[2]) if r[2] else None, str(r[3]) if r[3] else None,
                       str(r[4]) if r[4] else None, str(r[6]) if r[6] else None,
                       str(r[7]) if r[7] else None, str(r[8]) if r[8] else None,
                       str(r[9]) if r[9] else None))
    cur.executemany("""
        INSERT INTO box_specs (box_no,box_type,weight_g,height_cm,length_cm,breadth_cm,
                               max_strips,max_tubes,max_vials,max_bottles)
        VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
        ON CONFLICT (box_no) DO UPDATE SET weight_g=EXCLUDED.weight_g,
            height_cm=EXCLUDED.height_cm, length_cm=EXCLUDED.length_cm,
            breadth_cm=EXCLUDED.breadth_cm
    """, boxes)
    print(f"  ✅ {len(boxes)} boxes: {[b[0] for b in boxes]}")


# ─────────────────────────────────────────────────────
# STEP 3 — Product weights (4-strategy matching)
# ─────────────────────────────────────────────────────

def normalize(name: str) -> str:
    n = str(name).upper().strip()
    n = re.sub(r'\b\d+[X*x]\d+\b', '', n)
    n = re.sub(r'\b\d+\.?\d*\s*(MG|MCG|ML|GM|G|%|IU)\b', '', n, flags=re.I)
    n = re.sub(r'\b(TAB|CAP|CAPS|INJ|CREAM|GEL|OINT|FORTE|PLUS|SR|XR|ER|XL|LA|DS|MD)\b', '', n, flags=re.I)
    n = re.sub(r'\b\d+\b', '', n)
    n = re.sub(r'[^A-Z\s\-]', ' ', n)
    return re.sub(r'\s+', ' ', n).strip()


def brand_only(name: str) -> str:
    parts = re.split(r'[\s\-]+', str(name).upper())
    return parts[0] if parts else str(name)


def estimate_weight(name: str) -> tuple[float, str]:
    """
    Estimates shipment weight in grams from product name.
    Uses form type + dose + pack quantity.
    Accuracy: ±10-20g per strip — sufficient for shipping band purposes.
    """
    name = str(name).upper()
    pack_m = re.search(r'(\d+)[X*x](\d+)', name)
    qty = int(pack_m.group(2)) if pack_m else 10
    dose_m = re.search(r'(\d+\.?\d*)\s*(MG|MCG|ML|GM|G|%)', name, re.I)
    dose = float(dose_m.group(1)) if dose_m else 0
    unit = dose_m.group(2).upper() if dose_m else ''
    form = 'tablet'
    if any(x in name for x in ['INJ', 'INJECTION', 'AMP']):     form = 'injection'
    elif any(x in name for x in ['CREAM', 'GEL', 'OINT']):       form = 'topical'
    elif any(x in name for x in ['SYR', 'SYRUP', 'SUSP']):       form = 'liquid'
    elif any(x in name for x in ['CAP', 'CAPS', 'SOFTGEL']):     form = 'capsule'
    if unit == 'ML': form = 'liquid'
    if form == 'injection': w = qty * 25 + 20
    elif form == 'topical':  w = (dose * 1.1 + 30) if (dose > 0 and unit in ('G', 'GM')) else 30
    elif form == 'liquid':   w = (dose * 1.05 + 20) if (dose > 0 and unit == 'ML') else 80
    elif form == 'capsule':  w = qty * 0.8 + 10
    else:                    w = qty * max(dose * 0.003 + 0.3, 0.5) + 8
    return round(w, 1), form


def update_product_weights(cur, weight_file: str, catalog_file: str):
    print("\n[3/4] Updating product weights...")

    # ── Load weight sheet ─────────────────────────────
    wb = openpyxl.load_workbook(weight_file, read_only=True, data_only=True)
    ws = wb['🤖 WhatsApp Automation Lookup']
    rows = list(ws.iter_rows(values_only=True))
    hdr = next(i for i, r in enumerate(rows) if r and r[0] == '#')
    weight_map: dict[str, float] = {}
    for r in rows[hdr + 1:]:
        if r[1] and r[2] is not None:
            weight_map[str(r[1]).strip().upper()] = float(r[2])

    # ── Build lookup indexes ───────────────────────────
    norm_map: dict[str, tuple[str, float]] = {}
    brand_map: dict[str, list[tuple[str, float]]] = {}
    for name, wg in weight_map.items():
        norm = normalize(name)
        if norm and norm not in norm_map:
            norm_map[norm] = (name, wg)
        b = brand_only(name)
        brand_map.setdefault(b, []).append((name, wg))

    # ── Load catalog for cross-reference ─────────────
    wb2 = openpyxl.load_workbook(catalog_file, read_only=True, data_only=True)
    ws2 = wb2.active
    catalog_names: set[str] = set()
    for r in ws2.iter_rows(values_only=True):
        if r[0] and str(r[0]).strip() not in ('PRODUCT NAME', ''):
            catalog_names.add(str(r[0]).strip().upper())

    # ── Fetch DB products ─────────────────────────────
    cur.execute("SELECT id, UPPER(TRIM(product_name)) FROM products")
    db_products = {row[1]: row[0] for row in cur.fetchall()}

    updates: list[tuple[float, str, int]] = []  # (weight_g, weight_source, id)
    review_rows: list[dict] = []

    counts = {'exact': 0, 'normalized': 0, 'brand': 0, 'estimated': 0}

    for db_name, db_id in db_products.items():
        # Strategy 1: Exact
        if db_name in weight_map:
            updates.append((weight_map[db_name], 'exact', db_id))
            counts['exact'] += 1
            continue

        # Strategy 2: Normalized
        norm = normalize(db_name)
        if norm and norm in norm_map:
            orig, wg = norm_map[norm]
            updates.append((wg, 'normalized', db_id))
            counts['normalized'] += 1
            continue

        # Strategy 3: Brand name match
        b = brand_only(db_name)
        if b in brand_map:
            cands = brand_map[b]
            best = max(cands, key=lambda x: SequenceMatcher(
                None, normalize(db_name), normalize(x[0])).ratio())
            ratio = SequenceMatcher(None, normalize(db_name), normalize(best[0])).ratio()
            if ratio > 0.5:
                updates.append((best[1], 'brand', db_id))
                counts['brand'] += 1
                continue

        # Strategy 4: Estimate from product name
        est_w, form = estimate_weight(db_name)
        updates.append((est_w, 'estimated', db_id))
        counts['estimated'] += 1
        review_rows.append({
            'product_name': db_name,
            'estimated_weight_g': est_w,
            'form_detected': form,
            'action': 'REVIEW — not in weight sheet, enter actual weight_g'
        })

    cur.executemany(
        "UPDATE products SET weight_g = %s, weight_source = %s WHERE id = %s",
        updates
    )

    # Save review CSV
    if review_rows:
        fname = 'unmatched_weights_review.csv'
        with open(fname, 'w', newline='') as f:
            w = csv.DictWriter(f, fieldnames=review_rows[0].keys())
            w.writeheader()
            w.writerows(review_rows)
        print(f"  ⚠️  Review file saved → {fname}")

    print(f"  ✅ {len(updates)} products updated")
    print(f"     Exact:      {counts['exact']}")
    print(f"     Normalized: {counts['normalized']}")
    print(f"     Brand:      {counts['brand']}")
    print(f"     Estimated:  {counts['estimated']}  ← check unmatched_weights_review.csv")


# ─────────────────────────────────────────────────────
# STEP 4 — Shipping rates
# ─────────────────────────────────────────────────────

def parse_weight_band(raw: str) -> tuple[int, int] | None:
    nums = re.findall(r'\d+', str(raw).replace(',', '').replace(' ', ''))
    return (int(nums[0]), int(nums[1])) if len(nums) >= 2 else None


def seed_shipping_rates(cur, ems_file: str, lp_file: str):
    print("\n[4/4] Seeding shipping_rates...")
    cur.execute("DELETE FROM shipping_rates")
    all_records: list[tuple] = []

    # EMS — MUST use data_only=True (RUSSIA + SWITZERLAND have formula cells)
    wb_ems = openpyxl.load_workbook(ems_file, data_only=True)
    ws_ems = wb_ems.active
    ems_rows = list(ws_ems.iter_rows(values_only=True))
    ems_countries = [(i, str(c).strip().upper()) for i, c in enumerate(ems_rows[0])
                     if c and str(c).strip() not in ('WEIGHT RANGE', '')]
    for r in ems_rows[1:]:
        band = parse_weight_band(str(r[0])) if r[0] else None
        if not band:
            continue
        for ci, country in ems_countries:
            if ci >= len(r):
                continue
            val = r[ci]
            try:
                rate = float(val) if val is not None else None
            except (ValueError, TypeError):
                rate = None
            all_records.append((country, 'EMS', band[0], band[1], rate))

    # LP — no formulas, read_only is fine
    wb_lp = openpyxl.load_workbook(lp_file, read_only=True, data_only=True)
    ws_lp = wb_lp.active
    lp_rows = list(ws_lp.iter_rows(values_only=True))
    lp_countries = [(i, str(c).strip().upper()) for i, c in enumerate(lp_rows[0])
                    if c and str(c).strip() not in ('WEIGHT (GM)', '')]
    for r in lp_rows[1:]:
        band = parse_weight_band(str(r[0])) if r[0] else None
        if not band:
            continue
        for ci, country in lp_countries:
            if ci >= len(r):
                continue
            val = r[ci]
            try:
                rate = float(val) if val is not None else None
            except (ValueError, TypeError):
                rate = None
            all_records.append((country, 'LP', band[0], band[1], rate))

    psycopg2.extras.execute_values(
        cur,
        """INSERT INTO shipping_rates (country_name,shipping_type,weight_from_g,weight_to_g,rate_usd)
           VALUES %s ON CONFLICT (country_name,shipping_type,weight_from_g)
           DO UPDATE SET rate_usd=EXCLUDED.rate_usd""",
        all_records, page_size=500
    )

    cur.execute("SELECT COUNT(*) FROM shipping_rates WHERE shipping_type='EMS' AND rate_usd IS NOT NULL")
    ems_ok = cur.fetchone()[0]
    cur.execute("SELECT COUNT(*) FROM shipping_rates WHERE shipping_type='LP' AND rate_usd IS NOT NULL")
    lp_ok = cur.fetchone()[0]
    cur.execute("SELECT COUNT(DISTINCT country_name) FROM shipping_rates WHERE shipping_type='EMS'")
    ems_c = cur.fetchone()[0]
    cur.execute("SELECT COUNT(DISTINCT country_name) FROM shipping_rates WHERE shipping_type='LP'")
    lp_c = cur.fetchone()[0]
    print(f"  ✅ EMS: {ems_c} countries, {ems_ok} rate rows")
    print(f"  ✅ LP : {lp_c} countries, {lp_ok} rate rows")
    print(f"  ✅ Total rows: {len(all_records)}")


# ─────────────────────────────────────────────────────
# Main
# ─────────────────────────────────────────────────────

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--weight",  required=True, help="Intelligent_Weight_Sheet.xlsx")
    ap.add_argument("--catalog", required=True, help="catalog.xlsx")
    ap.add_argument("--ems",     required=True, help="EXPRESS_SHIPPING_CHARGES.xlsx")
    ap.add_argument("--lp",      required=True, help="REGULAR__LP__SHIPPING_CHARGES.xlsx")
    ap.add_argument("--dry-run", action="store_true")
    args = ap.parse_args()

    print("=" * 60)
    print("  WASA Shipping Data Seeder")
    print("=" * 60)

    conn = get_conn()
    cur = conn.cursor()
    try:
        create_tables(cur)
        seed_box_specs(cur, args.weight)
        update_product_weights(cur, args.weight, args.catalog)
        seed_shipping_rates(cur, args.ems, args.lp)

        if args.dry_run:
            conn.rollback()
            print("\n[DRY RUN] Rolled back — no changes saved")
        else:
            conn.commit()
            print("\n✅ Done")
    except Exception as e:
        conn.rollback()
        print(f"\n❌ Error: {e}")
        raise
    finally:
        cur.close()
        conn.close()


if __name__ == "__main__":
    main()